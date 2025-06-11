import torch
import os
from sklearn.model_selection import StratifiedKFold
import numpy as np
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

import logging
import sys
import builtins

logging.basicConfig(
    format="%(asctime)s | %(levelname)s : %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)


def matthews_correlation(y_true, y_pred):
    """Calculates the Matthews correlation coefficient measure for quality of binary classification problems."""
    y_pred = torch.tensor(y_pred, dtype=torch.float32)
    y_true = torch.tensor(y_true, dtype=torch.float32)

    y_pred_pos = torch.round(torch.clamp(y_pred, 0, 1))
    y_pred_neg = 1 - y_pred_pos

    y_pos = torch.round(torch.clamp(y_true, 0, 1))
    y_neg = 1 - y_pos

    tp = torch.sum(y_pos * y_pred_pos)
    tn = torch.sum(y_neg * y_pred_neg)

    fp = torch.sum(y_neg * y_pred_pos)
    fn = torch.sum(y_pos * y_pred_neg)

    numerator = tp * tn - fp * fn
    denominator = torch.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))

    return numerator / (denominator + torch.finfo(torch.float32).eps)


def mcc(tp, tn, fp, fn):
    sup = tp * tn - fp * fn
    inf = (tp + fp) * (tp + fn) * (tn + fp) * (tn + fn)
    if inf == 0:
        return 0
    else:
        return sup / np.sqrt(inf)


def eval_mcc(y_true, y_prob, show=False):
    """
    A fast implementation of Anokas mcc optimization code.

    This code takes as input probabilities, and selects the threshold that
    yields the best MCC score. It is efficient enough to be used as a
    custom evaluation function in xgboost

    Source: https://www.kaggle.com/cpmpml/optimizing-probabilities-for-best-mcc
    Source: https://www.kaggle.com/c/bosch-production-line-performance/forums/t/22917/optimising-probabilities-binary-prediction-script
    Creator: CPMP
    """
    idx = np.argsort(y_prob)
    y_true_sort = y_true[idx]
    n = y_true.shape[0]
    nump = 1.0 * np.sum(y_true)  # number of positive
    numn = n - nump  # number of negative
    tp = nump
    tn = 0.0
    fp = numn
    fn = 0.0
    best_mcc = 0.0
    best_id = -1
    prev_proba = -1
    best_proba = -1
    mccs = np.zeros(n)
    for i in range(n):
        # all items with idx < i are predicted negative while others are predicted positive
        # only evaluate mcc when probability changes
        proba = y_prob[idx[i]]
        if proba != prev_proba:
            new_mcc = mcc(tp, tn, fp, fn)
            if new_mcc >= best_mcc:
                best_mcc = new_mcc
                best_id = i
                best_proba = (prev_proba + proba) / 2.0 if prev_proba >= 0 else proba
            prev_proba = proba
        mccs[i] = new_mcc
        if y_true_sort[i] == 1:
            tp -= 1.0
            fn += 1.0
        else:
            fp -= 1.0
            tn += 1.0
    if show:
        y_pred = (y_prob >= best_proba).astype(int)
        score = matthews_correlation(y_true, y_pred)
        plt.plot(mccs)
        return best_proba, best_mcc, y_pred
    else:
        return best_proba, best_mcc, None


def create_dirs(dirs):
    """
    Input:
        dirs: a list of directories to create, in case these directories are not found
    Returns:
        exit_code: 0 if success, -1 if failure
    """
    try:
        for dir_ in dirs:
            if not os.path.exists(dir_):
                os.makedirs(dir_)
        return 0
    except Exception as err:
        print("Creating directories error: {0}".format(err))
        exit(-1)


def stratified_train_val_test_split(X, y, num_folds=5, seed=42):
    np.random.seed(seed)
    skf = StratifiedKFold(n_splits=num_folds, shuffle=True, random_state=seed)
    # get the indices of the folds
    splits = np.zeros_like(y, dtype=int)
    for fold_id, (_, val_idx) in enumerate(skf.split(X, y)):
        splits[val_idx] = fold_id

    fold_splits = []
    for val_fold in range(num_folds):
        test_fold = (val_fold + 1) % num_folds
        train_folds = [f for f in range(num_folds) if f not in [val_fold, test_fold]]

        train_idx = np.where(np.isin(splits, train_folds))[0]
        val_idx = np.where(splits == val_fold)[0]
        test_idx = np.where(splits == test_fold)[0]

        fold_splits.append((train_idx, val_idx, test_idx))

    return fold_splits


class Printer(object):
    """Class for printing output by refreshing the same line in the console, e.g. for indicating progress of a process"""

    def __init__(self, console=True):

        if console:
            self.print = self.dyn_print
        else:
            self.print = builtins.print

    @staticmethod
    def dyn_print(data):
        """Print things to stdout on one line, refreshing it dynamically"""
        sys.stdout.write("\r\x1b[K" + data.__str__())
        sys.stdout.flush()


def readable_time(time_difference):
    """Convert a float measuring time difference in seconds into a tuple of (hours, minutes, seconds)"""

    hours = time_difference // 3600
    minutes = (time_difference // 60) % 60
    seconds = time_difference % 60

    return hours, minutes, seconds


def export_performance_metrics(
    filepath, metrics_table, header, book=None, sheet_name="metrics"
):
    """Exports performance metrics on the validation set for all epochs to an excel file"""

    if os.path.exists(filepath):  # Create a records file for the first time
        book = load_workbook(filepath)

    if book is None:
        book = Workbook()  # new excel work book
        del book["Sheet"]  # remove default sheet

    book = write_table_to_sheet([header] + metrics_table, book, sheet_name=sheet_name)

    book.save(filepath)
    logger.info("Exported per epoch performance metrics in '{}'".format(filepath))

    return book


def write_table_to_sheet(table, work_book, sheet_name=None):
    """Writes a table implemented as a list of lists to an excel sheet in the given work book object"""

    sheet = work_book.create_sheet(sheet_name, index=0)

    for row_ind, row_list in enumerate(table):
        write_row(sheet, row_ind, row_list)

    return work_book


def write_row(sheet, row_ind, data_list):
    """Write a list to row_ind row of an excel sheet"""

    # row = sheet.row(row_ind)
    for col_ind, col_value in enumerate(data_list, start=1):
        sheet.cell(
            row=row_ind + 1, column=col_ind, value=col_value
        )  # row and col starts from 1 in openpyxl
    return


def save_model(path, epoch, model, optimizer=None):
    if isinstance(model, torch.nn.DataParallel):
        state_dict = model.module.state_dict()
    else:
        state_dict = model.state_dict()
    data = {"epoch": epoch, "state_dict": state_dict}
    if not (optimizer is None):
        data["optimizer"] = optimizer.state_dict()
    torch.save(data, path)


def export_record(filepath, values):
    """Adds a list of values as a bottom row of a table in a given excel file"""

    # read_book = xlrd.open_workbook(filepath, formatting_info=True)
    read_book = load_workbook(filepath)
    sheet = read_book.active
    last_row = sheet.max_row

    write_row(sheet, last_row, values)
    read_book.save(filepath)


def register_test_record(
    filepath,
    timestamp,
    experiment_name,
    best_metrics,
    final_metrics=None,
    test_metrics=None,
    comment="",
):
    """
    Adds the best and final metrics of a given experiment as a record in an excel sheet with other experiment records.
    Creates excel sheet if it doesn't exist.
    Args:
        filepath: path of excel file keeping records
        timestamp: string
        experiment_name: string
        best_metrics: dict of metrics at best epoch {metric_name: metric_value}. Includes "epoch" as first key
        final_metrics: dict of metrics at final epoch {metric_name: metric_value}. Includes "epoch" as first key
        test_metrics: dict of metrics at test stage {metric_name: metric_value}.
        comment: optional description
    """
    metrics_names, metrics_values = zip(*best_metrics.items())
    row_values = [timestamp, experiment_name, comment] + list(metrics_values)
    if final_metrics is not None:
        final_metrics_names, final_metrics_values = zip(*final_metrics.items())
        row_values += list(final_metrics_values)

    if test_metrics is not None:
        test_metrics_names, test_metrics_values = zip(*test_metrics.items())
        row_values += list(test_metrics_values)

    if not os.path.exists(filepath):  # Create a records file for the first time
        logger.warning(
            "Records file '{}' does not exist! Creating new file ...".format(filepath)
        )
        directory = os.path.dirname(filepath)
        if len(directory) and not os.path.exists(directory):
            os.makedirs(directory)
        header = ["Timestamp", "Name", "Comment"] + ["Best " + m for m in metrics_names]
        if final_metrics is not None:
            header += ["Final " + m for m in final_metrics_names]
        if test_metrics is not None:
            header += [m for m in test_metrics_names]
        book = Workbook()  # new excel work book
        del book["Sheet"]  # excel work book
        book = write_table_to_sheet([header, row_values], book, sheet_name="records")
        book.save(filepath)
    else:
        try:
            export_record(filepath, row_values)
        except Exception as x:
            alt_path = os.path.join(
                os.path.dirname(filepath), "record_" + experiment_name
            )
            logger.error(
                "Failed saving in: '{}'! Will save here instead: {}".format(
                    filepath, alt_path
                )
            )
            export_record(alt_path, row_values)
            filepath = alt_path

    logger.info("Exported performance record to '{}'".format(filepath))
